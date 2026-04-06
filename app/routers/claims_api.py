"""
API routes for NHIS Claims management.
"""
from fastapi import APIRouter, Depends, HTTPException, Request, Query, Form
from app.core.templates import templates
from fastapi.responses import RedirectResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime

from app.db.database import get_db
from app.core.deps import get_current_user, role_required
from app.models.claims_models import NHISClaim, ClaimStatus
from app.models.encounter_models import Encounter
from app.crud import claims_crud

router = APIRouter(tags=["NHIS Claims"])
# Register the age filter
from datetime import date
def calculate_age(dob):
    if not dob:
        return None
    if isinstance(dob, str):
        dob = date.fromisoformat(dob)
    today = date.today()
    age = today.year - dob.year
    if (today.month, today.day) < (dob.month, dob.day):
        age -= 1
    return age
templates.env.filters["age"] = calculate_age


@router.get("/claims", name="claims_dashboard")
def claims_dashboard(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    status_filter: Optional[str] = Query(None),
    provider_filter: Optional[str] = Query(None)  # "nhis" or provider name
):
    """
    Unified Claims Dashboard - Shows both NHIS and Private Insurance claims
    with provider badges and submission tracking
    """
    from app.models.billing_models import Invoice, InvoiceStatus, PaymentMethod
    from sqlalchemy.orm import joinedload
    from sqlalchemy import func
    from decimal import Decimal
    
    # ========== NHIS CLAIMS ==========
    nhis_query = db.query(NHISClaim).filter(NHISClaim.is_active == True)
    
    if status_filter and provider_filter != "private_insurance":
        # Normalize to lowercase to match database enum values
        status_filter = status_filter.lower()
        try:
            status_enum = ClaimStatus(status_filter)
            nhis_query = nhis_query.filter(NHISClaim.status == status_enum.value)
        except ValueError:
            pass
    
    # Filter by provider
    if provider_filter == "nhis" or not provider_filter:
        nhis_claims = nhis_query.options(
            joinedload(NHISClaim.patient),
            joinedload(NHISClaim.encounter),
            joinedload(NHISClaim.invoice)
        ).order_by(NHISClaim.claim_date.desc()).all()
    else:
        nhis_claims = []
    
    # NHIS Statistics
    total_nhis_claims = db.query(NHISClaim).filter(NHISClaim.is_active == True).count()
    submitted_nhis_claims = db.query(NHISClaim).filter(
        NHISClaim.status.in_([ClaimStatus.SUBMITTED.value, ClaimStatus.PROCESSING.value, ClaimStatus.APPROVED.value, ClaimStatus.PAID.value]),
        NHISClaim.is_active == True
    ).count()
    pending_nhis_claims = db.query(NHISClaim).filter(
        NHISClaim.status == ClaimStatus.PENDING.value,
        NHISClaim.is_active == True
    ).count()
    approved_nhis_claims = db.query(NHISClaim).filter(
        NHISClaim.status == ClaimStatus.APPROVED.value,
        NHISClaim.is_active == True
    ).count()
    paid_nhis_claims = db.query(NHISClaim).filter(
        NHISClaim.status == ClaimStatus.PAID.value,
        NHISClaim.is_active == True
    ).count()
    
    # NHIS Financial Summary
    nhis_total_amount = db.query(func.sum(NHISClaim.total_amount)).filter(
        NHISClaim.is_active == True
    ).scalar() or Decimal('0.00')
    nhis_submitted_amount = db.query(func.sum(NHISClaim.total_amount)).filter(
        NHISClaim.status.in_([ClaimStatus.SUBMITTED.value, ClaimStatus.PROCESSING.value, ClaimStatus.APPROVED.value, ClaimStatus.PAID.value]),
        NHISClaim.is_active == True
    ).scalar() or Decimal('0.00')
    nhis_paid_amount = db.query(func.sum(NHISClaim.approved_amount)).filter(
        NHISClaim.status == ClaimStatus.PAID.value,
        NHISClaim.is_active == True
    ).scalar() or Decimal('0.00')
    
    # ========== PRIVATE INSURANCE CLAIMS ==========
    private_query = db.query(Invoice).filter(
        Invoice.is_active == True,
        Invoice.payment_mechanism == PaymentMethod.PRIVATE_INSURANCE.value
    )
    
    if status_filter and provider_filter != "nhis":
        try:
            status_enum = InvoiceStatus(status_filter)
            private_query = private_query.filter(Invoice.status == status_enum.value)
        except ValueError:
            pass
    
    # Filter by provider
    if provider_filter and provider_filter != "nhis" and provider_filter != "private_insurance":
        from app.models.patient_models import Patient
        private_query = private_query.join(Patient).filter(
            func.coalesce(Invoice.insurance_provider, Patient.insurance_provider) == provider_filter
        )
    
    if provider_filter != "nhis":
        private_invoices = private_query.options(
            joinedload(Invoice.patient),
            joinedload(Invoice.encounter)
        ).order_by(Invoice.invoice_date.desc()).all()
    else:
        private_invoices = []
    
    # Group private insurance by provider
    invoices_by_provider = {}
    total_amount_by_provider = {}
    submitted_count_by_provider = {}
    
    for invoice in private_invoices:
        provider = invoice.insurance_provider or invoice.patient.insurance_provider or "Unknown Provider"
        
        if provider not in invoices_by_provider:
            invoices_by_provider[provider] = []
            total_amount_by_provider[provider] = Decimal('0.00')
            submitted_count_by_provider[provider] = 0
        
        invoices_by_provider[provider].append(invoice)
        total_amount_by_provider[provider] += invoice.total_amount
        
        # Count submitted (pending, partially_paid, paid are considered submitted)
        if invoice.status.value in ['pending', 'partially_paid', 'paid']:
            submitted_count_by_provider[provider] += 1
    
    # Private Insurance Statistics
    total_private_invoices = len(private_invoices)
    pending_private_invoices = sum(1 for inv in private_invoices if inv.status == InvoiceStatus.PENDING.value)
    paid_private_invoices = sum(1 for inv in private_invoices if inv.status == InvoiceStatus.PAID.value)
    
    total_private_amount = sum(inv.total_amount for inv in private_invoices)
    total_pending_private_amount = sum(inv.balance for inv in private_invoices if inv.status == InvoiceStatus.PENDING.value)
    total_paid_private_amount = sum(inv.total_amount for inv in private_invoices if inv.status == InvoiceStatus.PAID.value)
    
    # Get list of unique insurance providers
    insurance_providers = sorted(invoices_by_provider.keys())
    
    context = {
        "request": request,
        "title": "Insurance Claims Dashboard",
        "current_user": current_user,
        "user_role": current_user.role.name,
        # NHIS Data
        "nhis_claims": nhis_claims,
        "total_nhis_claims": total_nhis_claims,
        "submitted_nhis_claims": submitted_nhis_claims,
        "pending_nhis_claims": pending_nhis_claims,
        "approved_nhis_claims": approved_nhis_claims,
        "paid_nhis_claims": paid_nhis_claims,
        "nhis_total_amount": nhis_total_amount,
        "nhis_submitted_amount": nhis_submitted_amount,
        "nhis_paid_amount": nhis_paid_amount,
        # Private Insurance Data
        "private_invoices": private_invoices,
        "invoices_by_provider": invoices_by_provider,
        "total_amount_by_provider": total_amount_by_provider,
        "submitted_count_by_provider": submitted_count_by_provider,
        "insurance_providers": insurance_providers,
        "total_private_invoices": total_private_invoices,
        "pending_private_invoices": pending_private_invoices,
        "paid_private_invoices": paid_private_invoices,
        "total_private_amount": total_private_amount,
        "total_pending_private_amount": total_pending_private_amount,
        "total_paid_private_amount": total_paid_private_amount,
        # Filters
        "status_filter": status_filter,
        "provider_filter": provider_filter,
    }
    return templates.TemplateResponse("claims/unified_dashboard.html", context)


@router.get("/claims/private-insurance", name="private_insurance_claims_dashboard")
def private_insurance_claims_dashboard(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    status_filter: Optional[str] = Query(None),
    insurance_provider: Optional[str] = Query(None)
):
    """Private Insurance Claims dashboard - aggregates all bills and claims grouped by insurance company"""
    from app.models.billing_models import Invoice, InvoiceStatus, PaymentMethod
    from sqlalchemy.orm import joinedload
    from sqlalchemy import func
    from decimal import Decimal
    
    # Base query for private insurance invoices
    base_query = db.query(Invoice).filter(
        Invoice.is_active == True,
        Invoice.payment_mechanism == PaymentMethod.PRIVATE_INSURANCE.value
    )
    
    if status_filter:
        # Normalize to lowercase to match database enum values
        status_filter = status_filter.lower()
        try:
            status_enum = InvoiceStatus(status_filter)
            base_query = base_query.filter(Invoice.status == status_enum.value)
        except ValueError:
            pass
    
    # Get all invoices
    all_invoices = base_query.options(
        joinedload(Invoice.patient),
        joinedload(Invoice.encounter)
    ).order_by(Invoice.invoice_date.desc()).all()
    
    # Group invoices by insurance provider
    invoices_by_provider = {}
    total_amount_by_provider = {}
    pending_amount_by_provider = {}
    paid_amount_by_provider = {}
    
    for invoice in all_invoices:
        # Get insurance provider from invoice or patient
        provider = invoice.insurance_provider or invoice.patient.insurance_provider or "Unknown Provider"
        
        if provider not in invoices_by_provider:
            invoices_by_provider[provider] = []
            total_amount_by_provider[provider] = Decimal('0.00')
            pending_amount_by_provider[provider] = Decimal('0.00')
            paid_amount_by_provider[provider] = Decimal('0.00')
        
        invoices_by_provider[provider].append(invoice)
        total_amount_by_provider[provider] += invoice.total_amount
        
        if invoice.status == InvoiceStatus.PENDING.value:
            pending_amount_by_provider[provider] += invoice.balance
        elif invoice.status == InvoiceStatus.PAID.value:
            paid_amount_by_provider[provider] += invoice.total_amount
    
    # Filter by insurance provider if specified
    if insurance_provider:
        if insurance_provider in invoices_by_provider:
            filtered_invoices = invoices_by_provider[insurance_provider]
        else:
            filtered_invoices = []
    else:
        filtered_invoices = all_invoices[:100]  # Limit display
    
    # Statistics
    total_invoices = len(all_invoices)
    pending_invoices = sum(1 for inv in all_invoices if inv.status == InvoiceStatus.PENDING.value)
    paid_invoices = sum(1 for inv in all_invoices if inv.status == InvoiceStatus.PAID.value)
    
    total_amount = sum(inv.total_amount for inv in all_invoices)
    total_pending_amount = sum(inv.balance for inv in all_invoices if inv.status == InvoiceStatus.PENDING.value)
    total_paid_amount = sum(inv.total_amount for inv in all_invoices if inv.status == InvoiceStatus.PAID.value)
    
    # Get list of unique insurance providers
    insurance_providers = sorted(invoices_by_provider.keys())
    
    context = {
        "request": request,
        "title": "Private Insurance Claims",
        "current_user": current_user,
        "user_role": current_user.role.name,
        "invoices": filtered_invoices,
        "invoices_by_provider": invoices_by_provider,
        "total_amount_by_provider": total_amount_by_provider,
        "pending_amount_by_provider": pending_amount_by_provider,
        "paid_amount_by_provider": paid_amount_by_provider,
        "insurance_providers": insurance_providers,
        "status_filter": status_filter,
        "insurance_provider_filter": insurance_provider,
        "total_invoices": total_invoices,
        "pending_invoices": pending_invoices,
        "paid_invoices": paid_invoices,
        "total_amount": total_amount,
        "total_pending_amount": total_pending_amount,
        "total_paid_amount": total_paid_amount
    }
    return templates.TemplateResponse("claims/private_insurance_claims.html", context)


@router.post("/claims/create-from-encounter/{encounter_id}", name="create_claim_from_encounter", status_code=302)
def create_claim_from_encounter(
    request: Request,
    encounter_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"]))
):
    """Create an NHIS claim from an encounter"""
    try:
        claim = claims_crud.create_claim_from_encounter(db, encounter_id, current_user.id)
        return RedirectResponse(
            url=f"/claims/{claim.id}?status=created",
            status_code=302
        )
    except ValueError as e:
        return RedirectResponse(
            url=f"/encounters/{encounter_id}?error={str(e)}",
            status_code=302
        )


@router.get("/claims/export", name="export_nhis_claims")
def export_nhis_claims(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    status_filter: Optional[str] = Query(None),
    format: str = Query("csv", regex="^(csv|excel)$")
):
    """Export NHIS claims to CSV or Excel for submission to NHIA."""
    from fastapi.responses import StreamingResponse
    import csv
    import io
    from datetime import datetime
    from sqlalchemy.orm import joinedload
    
    query = db.query(NHISClaim).filter(NHISClaim.is_active == True)
    
    if status_filter:
        # Normalize to lowercase to match database enum values
        status_filter = status_filter.lower()
        try:
            status_enum = ClaimStatus(status_filter)
            query = query.filter(NHISClaim.status == status_enum.value)
        except ValueError:
            pass
    
    claims = query.options(
        joinedload(NHISClaim.patient),
        joinedload(NHISClaim.encounter),
        joinedload(NHISClaim.invoice)
    ).order_by(NHISClaim.claim_date.desc()).all()
    
    if format == "csv":
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Claim Number", "Claim Date", "Patient Name", "Patient ID", "NHIS Number",
            "Encounter ID", "Invoice Number", "Total Amount", "NHIS Amount", "Co-Pay Amount",
            "Status", "Submitted At", "Submission Reference", "Approved Amount", "Processed At"
        ])
        
        # Data rows
        for claim in claims:
            patient = claim.patient
            invoice = claim.invoice
            writer.writerow([
                claim.claim_number or f"CLAIM-{claim.id}",
                claim.claim_date.strftime('%Y-%m-%d') if claim.claim_date else '',
                f"{patient.first_name} {patient.last_name}" if patient else 'N/A',
                patient.id if patient else 'N/A',
                claim.nhis_number or (patient.nhis_number if patient else 'N/A'),
                claim.encounter_id if claim.encounter_id else 'N/A',
                invoice.invoice_number if invoice else 'N/A',
                str(claim.total_amount or 0),
                str(claim.nhis_amount or 0),
                str(claim.co_pay_amount or 0),
                claim.status.value if claim.status else 'N/A',
                claim.submitted_at.strftime('%Y-%m-%d %H:%M') if claim.submitted_at else '',
                claim.submission_reference or 'N/A',
                str(claim.approved_amount) if claim.approved_amount else 'N/A',
                claim.processed_at.strftime('%Y-%m-%d %H:%M') if claim.processed_at else ''
            ])
        
        output.seek(0)
        filename = f"nhis_claims_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "excel":
        # Generate Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "NHIS Claims"
            
            # Header row
            headers = [
                "Claim Number", "Claim Date", "Patient Name", "Patient ID", "NHIS Number",
                "Encounter ID", "Invoice Number", "Total Amount", "NHIS Amount", "Co-Pay Amount",
                "Status", "Submitted At", "Submission Reference", "Approved Amount", "Processed At"
            ]
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for claim in claims:
                patient = claim.patient
                invoice = claim.invoice
                ws.append([
                    claim.claim_number or f"CLAIM-{claim.id}",
                    claim.claim_date.strftime('%Y-%m-%d') if claim.claim_date else '',
                    f"{patient.first_name} {patient.last_name}" if patient else 'N/A',
                    patient.id if patient else 'N/A',
                    claim.nhis_number or (patient.nhis_number if patient else 'N/A'),
                    claim.encounter_id if claim.encounter_id else 'N/A',
                    invoice.invoice_number if invoice else 'N/A',
                    float(claim.total_amount or 0),
                    float(claim.nhis_amount or 0),
                    float(claim.co_pay_amount or 0),
                    claim.status.value if claim.status else 'N/A',
                    claim.submitted_at.strftime('%Y-%m-%d %H:%M') if claim.submitted_at else '',
                    claim.submission_reference or 'N/A',
                    float(claim.approved_amount) if claim.approved_amount else 0,
                    claim.processed_at.strftime('%Y-%m-%d %H:%M') if claim.processed_at else ''
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"nhis_claims_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl is required for Excel export. Install it with: pip install openpyxl")


@router.get("/claims/{claim_id}", name="view_claim")
def view_claim(
    request: Request,
    claim_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"]))
):
    """View a specific NHIS claim"""
    claim = claims_crud.get_claim(db, claim_id)
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    
    context = {
        "request": request,
        "title": f"NHIS Claim {claim.claim_number}",
        "current_user": current_user,
        "user_role": current_user.role.name,
        "claim": claim,
        "encounter": claim.encounter,
        "patient": claim.patient,
        "invoice": claim.invoice
    }
    return templates.TemplateResponse("claims/claim_detail.html", context)


@router.post("/claims/{claim_id}/submit", name="submit_claim", status_code=302)
def submit_claim(
    request: Request,
    claim_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"]))
):
    """
    Submit claim to NHIA.
    Note: Actual API submission is pending NHIA API integration.
    This currently marks the claim as submitted.
    """
    claim = claims_crud.get_claim(db, claim_id)
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    
    # TODO: When NHIA API is available, implement actual submission
    # For now, just update status
    updated_claim = claims_crud.update_claim_status(
        db, claim_id, ClaimStatus.SUBMITTED
    )
    
    if updated_claim:
        # Set submission reference
        updated_claim.submission_reference = f"SUBM-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        db.commit()
    else:
        raise HTTPException(status_code=400, detail="Failed to submit claim")
    
    return RedirectResponse(
        url=f"/claims/{claim_id}?status=submitted"
    )


@router.post("/claims/{claim_id}/update-status", name="update_nhis_claim_status")
def update_nhis_claim_status(
    request: Request,
    claim_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    new_status: str = Form(...),
    notes: Optional[str] = Form(None)
):
    """
    Update the status of an NHIS claim.
    Uses the update_claim_status function which automatically handles:
    - Invoice status updates
    - Payment record creation (when status = PAID)
    - OPD visit payment status sync
    """
    try:
        status_enum = ClaimStatus(new_status)
        
        # Use the update_claim_status function which handles automatic payment updates
        updated_claim = claims_crud.update_claim_status(
            db=db,
            claim_id=claim_id,
            new_status=status_enum,
            rejection_reason=notes if status_enum == ClaimStatus.REJECTED else None
        )
        
        if not updated_claim:
            raise HTTPException(status_code=404, detail="Claim not found")
        
        # Add notes if provided
        if notes:
            updated_claim.notes = (updated_claim.notes or '') + f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {notes}"
            db.commit()
            db.refresh(updated_claim)
        
        return RedirectResponse(
            url=f"/claims?status=updated&claim_id={claim_id}",
            status_code=status.HTTP_302_FOUND
        )
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid status: {new_status}")


@router.get("/claims/private-insurance/export", name="export_private_insurance_invoices")
def export_private_insurance_invoices(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    status_filter: Optional[str] = Query(None),
    insurance_provider: Optional[str] = Query(None),
    format: str = Query("csv", regex="^(csv|excel)$")
):
    """Export private insurance invoices to CSV or Excel for submission to providers."""
    from app.models.billing_models import Invoice, InvoiceStatus, PaymentMethod
    from sqlalchemy.orm import joinedload
    from fastapi.responses import StreamingResponse
    import csv
    import io
    from datetime import datetime
    
    # Base query for private insurance invoices
    base_query = db.query(Invoice).filter(
        Invoice.is_active == True,
        Invoice.payment_mechanism == PaymentMethod.PRIVATE_INSURANCE.value
    ).options(
        joinedload(Invoice.patient),
        joinedload(Invoice.encounter)
    )
    
    if status_filter:
        # Normalize to lowercase to match database enum values
        status_filter = status_filter.lower()
        try:
            status_enum = InvoiceStatus(status_filter)
            base_query = base_query.filter(Invoice.status == status_enum.value)
        except ValueError:
            pass
    
    if insurance_provider:
        from app.models.patient_models import Patient
        base_query = base_query.join(Patient).filter(
            Patient.insurance_provider == insurance_provider
        )
    
    invoices = base_query.order_by(Invoice.invoice_date.desc()).all()
    
    if format == "csv":
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Invoice Number", "Invoice Date", "Patient Name", "Patient ID", 
            "Insurance Provider", "Policy Number", "Encounter ID",
            "Total Amount", "Paid Amount", "Balance", "Status", "Due Date"
        ])
        
        # Data rows
        for invoice in invoices:
            patient = invoice.patient
            writer.writerow([
                invoice.invoice_number or f"INV-{invoice.id}",
                invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
                f"{patient.first_name} {patient.last_name}" if patient else 'N/A',
                patient.id if patient else 'N/A',
                invoice.insurance_provider or (patient.insurance_provider if patient else 'N/A'),
                invoice.insurance_policy_number or (patient.insurance_policy_number if patient else 'N/A'),
                invoice.encounter_id if invoice.encounter_id else 'N/A',
                str(invoice.total_amount or 0),
                str(invoice.paid_amount or 0),
                str(invoice.balance or 0),
                invoice.status.value if invoice.status else 'N/A',
                invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else ''
            ])
        
        output.seek(0)
        filename = f"private_insurance_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "excel":
        # Generate Excel - check if openpyxl is available
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError as e:
            raise HTTPException(
                status_code=500, 
                detail=f"openpyxl is required for Excel export. Install it with: pip install openpyxl. Error: {str(e)}"
            )
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Private Insurance Invoices"
            
            # Header row
            headers = [
                "Invoice Number", "Invoice Date", "Patient Name", "Patient ID", 
                "Insurance Provider", "Policy Number", "Encounter ID",
                "Total Amount", "Paid Amount", "Balance", "Status", "Due Date"
            ]
            ws.append(headers)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for invoice in invoices:
                patient = invoice.patient
                ws.append([
                    invoice.invoice_number or f"INV-{invoice.id}",
                    invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
                    f"{patient.first_name} {patient.last_name}" if patient else 'N/A',
                    patient.id if patient else 'N/A',
                    invoice.insurance_provider or (patient.insurance_provider if patient else 'N/A'),
                    invoice.insurance_policy_number or (patient.insurance_policy_number if patient else 'N/A'),
                    invoice.encounter_id if invoice.encounter_id else 'N/A',
                    float(invoice.total_amount or 0),
                    float(invoice.paid_amount or 0),
                    float(invoice.balance or 0),
                    invoice.status.value if invoice.status else 'N/A',
                    invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else ''
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"private_insurance_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error generating Excel file: {str(e)}"
            )


@router.post("/claims/private-insurance/invoices/{invoice_id}/update-status", name="update_private_insurance_invoice_status")
def update_private_insurance_invoice_status(
    request: Request,
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    new_status: str = Form(...),
    notes: Optional[str] = Form(None)
):
    """Update the status of a private insurance invoice."""
    from app.models.billing_models import Invoice, InvoiceStatus, PaymentMethod
    
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.is_active == True,
        Invoice.payment_mechanism == PaymentMethod.PRIVATE_INSURANCE.value
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    try:
        status_enum = InvoiceStatus(new_status)
        invoice.status = status_enum.value
        
        if notes:
            # Store notes in invoice notes field if available, or create a comment
            if hasattr(invoice, 'notes'):
                invoice.notes = (invoice.notes or '') + f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {notes}"
        
        db.commit()
        db.refresh(invoice)
        
        return RedirectResponse(
            url=f"/claims/private-insurance?status=updated&invoice_id={invoice_id}",
            status_code=status.HTTP_302_FOUND
        )
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid status: {new_status}")


@router.post("/claims/batch-update-nhis", name="batch_update_nhis_claims_status")
def batch_update_nhis_claims_status(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    claim_ids: str = Form(...),
    new_status: str = Form(...),
    notes: Optional[str] = Form(None)
):
    """
    Batch update status for multiple NHIS claims.
    Updates all selected claims to the new status and handles automatic payment updates.
    """
    try:
        status_enum = ClaimStatus(new_status)
        claim_id_list = [int(id.strip()) for id in claim_ids.split(',') if id.strip()]
        
        if not claim_id_list:
            raise HTTPException(status_code=400, detail="No claims selected")
        
        updated_count = 0
        failed_count = 0
        errors = []
        
        for claim_id in claim_id_list:
            try:
                updated_claim = claims_crud.update_claim_status(
                    db=db,
                    claim_id=claim_id,
                    new_status=status_enum,
                    rejection_reason=notes if status_enum == ClaimStatus.REJECTED else None
                )
                
                if updated_claim:
                    # Add notes if provided
                    if notes:
                        updated_claim.notes = (updated_claim.notes or '') + f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M')}] Batch Update: {notes}"
                        db.commit()
                        db.refresh(updated_claim)
                    updated_count += 1
                else:
                    failed_count += 1
                    errors.append(f"Claim {claim_id} not found")
            except Exception as e:
                failed_count += 1
                errors.append(f"Claim {claim_id}: {str(e)}")
        
        # Build redirect URL with results
        result_msg = f"Updated {updated_count} claim(s)"
        if failed_count > 0:
            result_msg += f", {failed_count} failed"
        
        return RedirectResponse(
            url=f"/claims?status=batch_updated&updated={updated_count}&failed={failed_count}",
            status_code=302
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid status: {new_status}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Batch update failed: {str(e)}")


@router.post("/claims/batch-update-private-insurance", name="batch_update_private_insurance_invoices_status")
def batch_update_private_insurance_invoices_status(
    request: Request,
    db: Session = Depends(get_db),
    current_user = Depends(role_required(["Admin", "Finance"])),
    invoice_ids: str = Form(...),
    new_status: str = Form(...),
    notes: Optional[str] = Form(None)
):
    """
    Batch update status for multiple private insurance invoices.
    """
    from app.models.billing_models import Invoice, InvoiceStatus, PaymentMethod
    
    try:
        status_enum = InvoiceStatus(new_status)
        invoice_id_list = [int(id.strip()) for id in invoice_ids.split(',') if id.strip()]
        
        if not invoice_id_list:
            raise HTTPException(status_code=400, detail="No invoices selected")
        
        updated_count = 0
        failed_count = 0
        
        for invoice_id in invoice_id_list:
            try:
                invoice = db.query(Invoice).filter(
                    Invoice.id == invoice_id,
                    Invoice.is_active == True,
                    Invoice.payment_mechanism == PaymentMethod.PRIVATE_INSURANCE.value
                ).first()
                
                if invoice:
                    invoice.status = status_enum.value
                    
                    if notes:
                        if hasattr(invoice, 'notes'):
                            invoice.notes = (invoice.notes or '') + f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M')}] Batch Update: {notes}"
                    
                    db.commit()
                    db.refresh(invoice)
                    updated_count += 1
                else:
                    failed_count += 1
            except Exception as e:
                failed_count += 1
                db.rollback()
        
        return RedirectResponse(
            url=f"/claims?status=batch_updated&updated={updated_count}&failed={failed_count}&type=private",
            status_code=302
        )
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid status: {new_status}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Batch update failed: {str(e)}")

