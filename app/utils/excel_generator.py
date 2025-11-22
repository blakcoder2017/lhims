"""
Excel Generation Utilities

Utilities for generating Excel reports using openpyxl.
"""
from typing import Dict, Any
from io import BytesIO
from datetime import datetime
from decimal import Decimal

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_income_statement_excel(context: Dict[str, Any]) -> bytes:
    """Generate Excel report for income statement."""
    if not OPENPYXL_AVAILABLE:
        # Return a simple CSV as fallback
        return _generate_income_statement_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    total_fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    total_fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="INCOME STATEMENT (PROFIT & LOSS)").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period
    period_label = context.get('period_label', '')
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=period_label)
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%B %d, %Y at %I:%M %p'))
    row += 2
    
    # Revenue Section
    ws.cell(row=row, column=1, value="REVENUE").font = subtitle_font
    row += 1
    headers = ['Description', 'Amount (GHS)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
    row += 1
    
    revenue_by_service = context.get('revenue_by_service', {})
    for service_type, amount in revenue_by_service.items():
        ws.cell(row=row, column=1, value=service_type.replace('_', ' ').title()).border = border
        ws.cell(row=row, column=2, value=float(amount)).border = border
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1
    
    total_revenue = float(context.get('total_revenue', Decimal('0.00')))
    ws.cell(row=row, column=1, value="TOTAL REVENUE").font = total_font
    ws.cell(row=row, column=1).fill = total_fill_green
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value=total_revenue).font = total_font
    ws.cell(row=row, column=2).fill = total_fill_green
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=2).border = border
    row += 2
    
    # Expenses Section
    ws.cell(row=row, column=1, value="EXPENSES").font = subtitle_font
    row += 1
    headers = ['Category', 'Amount (GHS)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
    row += 1
    
    expenses_by_category = context.get('expenses_by_category', {})
    for category, amount in expenses_by_category.items():
        ws.cell(row=row, column=1, value=category.replace('_', ' ').title()).border = border
        ws.cell(row=row, column=2, value=float(amount)).border = border
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1
    
    total_expenses = float(context.get('total_expenses', Decimal('0.00')))
    ws.cell(row=row, column=1, value="TOTAL EXPENSES").font = total_font
    ws.cell(row=row, column=1).fill = total_fill_red
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value=total_expenses).font = total_font
    ws.cell(row=row, column=2).fill = total_fill_red
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=2).border = border
    row += 2
    
    # Net Income Section
    net_income = float(context.get('net_income', Decimal('0.00')))
    profit_margin = float(context.get('profit_margin', Decimal('0.00')))
    net_income_fill = total_fill_green if net_income >= 0 else total_fill_red
    
    ws.cell(row=row, column=1, value="NET INCOME").font = total_font
    ws.cell(row=row, column=1).fill = net_income_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value=net_income).font = total_font
    ws.cell(row=row, column=2).fill = net_income_fill
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=2).border = border
    row += 1
    ws.cell(row=row, column=1, value="Profit Margin").font = total_font
    ws.cell(row=row, column=1).fill = net_income_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value=f"{profit_margin:.2f}%").font = total_font
    ws.cell(row=row, column=2).fill = net_income_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=2).border = border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_income_statement_csv(context: Dict[str, Any]) -> str:
    """Fallback CSV for income statement."""
    from app.utils.csv_generator import generate_income_statement_csv
    return generate_income_statement_csv(context)


def generate_opd_report_excel(context: Dict[str, Any]) -> bytes:
    """Generate detailed Excel report for OPD visits."""
    if not OPENPYXL_AVAILABLE:
        return _generate_opd_report_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "OPD Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:H{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value="OPD DETAILED REPORT").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period and summary
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context['start_date']} to {context['end_date']}")
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%Y-%m-%d %H:%M'))
    row += 1
    ws.cell(row=row, column=1, value="Total Visits:").font = subtitle_font
    ws.cell(row=row, column=2, value=context.get('total_visits', 0))
    row += 1
    ws.cell(row=row, column=1, value="Total Revenue:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"GHS {float(context.get('total_revenue', Decimal('0.00'))):,.2f}")
    row += 2
    
    # Payment Breakdown
    payment_breakdown = context.get('payment_breakdown', {})
    if payment_breakdown:
        ws.cell(row=row, column=1, value="PAYMENT BREAKDOWN").font = subtitle_font
        row += 1
        headers = ['Payment Status', 'Count', 'Total (GHS)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        for status, data in payment_breakdown.items():
            ws.cell(row=row, column=1, value=status.title()).border = border
            ws.cell(row=row, column=2, value=data.get('count', 0)).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=float(data.get('total', Decimal('0.00')))).border = border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Gender Distribution
    gender_breakdown = context.get('gender_breakdown', {})
    if gender_breakdown:
        ws.cell(row=row, column=1, value="GENDER DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Gender', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(gender_breakdown.values())
        for gender, count in gender_breakdown.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            ws.cell(row=row, column=1, value=gender or 'Unknown').border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Age Distribution
    age_groups = context.get('age_groups', {})
    if age_groups:
        ws.cell(row=row, column=1, value="AGE DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Age Group', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(age_groups.values())
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            age_label = f"{age_group} years" if age_group != "65+" else "65+ years"
            ws.cell(row=row, column=1, value=age_label).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Detailed Visits
    visits = context.get('visits', [])
    if visits:
        ws.cell(row=row, column=1, value="DETAILED VISITS").font = subtitle_font
        row += 1
        headers = ['OPD Number', 'Patient Name', 'Patient #', 'Visit Date', 'Status', 'Payment Status', 'Visit Type', 'Total Charges (GHS)', 'Encounters']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [3, 4, 5, 6, 7, 9] else 'left' if col in [1, 2] else 'right')
        row += 1
        
        for visit in visits:
            patient_name = f"{visit.patient.first_name} {visit.patient.last_name}" if visit.patient else "N/A"
            patient_number = visit.patient.patient_number if visit.patient else "N/A"
            visit_date = visit.visit_date.strftime('%Y-%m-%d') if visit.visit_date else 'N/A'
            status = visit.status.value.title() if hasattr(visit.status, 'value') else str(visit.status).title()
            payment_status = visit.payment_status.title() if visit.payment_status else 'N/A'
            visit_type = visit.visit_type.title() if visit.visit_type else 'N/A'
            total_charges = float(visit.total_charges) if hasattr(visit, 'total_charges') else 0.00
            encounters_count = len(visit.encounters) if visit.encounters else 0
            
            ws.cell(row=row, column=1, value=visit.opd_number).border = border
            ws.cell(row=row, column=2, value=patient_name).border = border
            ws.cell(row=row, column=3, value=patient_number).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=visit_date).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=status).border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=payment_status).border = border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=7, value=visit_type).border = border
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8, value=total_charges).border = border
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=9, value=encounters_count).border = border
            ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_ipd_report_excel(context: Dict[str, Any]) -> bytes:
    """Generate detailed Excel report for IPD admissions."""
    if not OPENPYXL_AVAILABLE:
        return _generate_ipd_report_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "IPD Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:J{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:J{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:J{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    # Title
    ws.merge_cells(f'A{row}:J{row}')
    ws.cell(row=row, column=1, value="IPD DETAILED REPORT").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period and summary
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context['start_date']} to {context['end_date']}")
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%Y-%m-%d %H:%M'))
    row += 1
    ws.cell(row=row, column=1, value="Total Admissions:").font = subtitle_font
    ws.cell(row=row, column=2, value=context.get('total_admissions', 0))
    row += 1
    ws.cell(row=row, column=1, value="Average LOS:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context.get('avg_los', 0):.1f} days")
    row += 1
    ws.cell(row=row, column=1, value="Total Revenue:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"GHS {float(context.get('total_revenue', Decimal('0.00'))):,.2f}")
    row += 2
    
    # Ward Breakdown
    ward_breakdown = context.get('ward_breakdown', {})
    if ward_breakdown:
        ws.cell(row=row, column=1, value="WARD BREAKDOWN").font = subtitle_font
        row += 1
        headers = ['Ward', 'Count', 'Avg LOS (days)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        for ward_name, data in ward_breakdown.items():
            avg_los = data['total_los'] / data['count'] if data['count'] > 0 else 0
            ws.cell(row=row, column=1, value=ward_name).border = border
            ws.cell(row=row, column=2, value=data['count']).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{avg_los:.1f}").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Gender Distribution
    gender_breakdown = context.get('gender_breakdown', {})
    if gender_breakdown:
        ws.cell(row=row, column=1, value="GENDER DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Gender', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(gender_breakdown.values())
        for gender, count in gender_breakdown.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            ws.cell(row=row, column=1, value=gender or 'Unknown').border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Age Distribution
    age_groups = context.get('age_groups', {})
    if age_groups:
        ws.cell(row=row, column=1, value="AGE DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Age Group', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(age_groups.values())
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            age_label = f"{age_group} years" if age_group != "65+" else "65+ years"
            ws.cell(row=row, column=1, value=age_label).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Detailed Admissions
    admissions = context.get('admissions', [])
    if admissions:
        ws.cell(row=row, column=1, value="DETAILED ADMISSIONS").font = subtitle_font
        row += 1
        headers = ['Admission Number', 'Patient Name', 'Patient #', 'Ward', 'Bed', 'Admission Date', 'Discharge Date', 'Status', 'LOS (days)', 'Revenue (GHS)', 'Encounters']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [3, 4, 5, 6, 7, 8, 9, 11] else 'left' if col in [1, 2] else 'right')
        row += 1
        
        admission_los = context.get('admission_los', {})
        for admission in admissions:
            patient_name = f"{admission.patient.first_name} {admission.patient.last_name}" if admission.patient else "N/A"
            patient_number = admission.patient.patient_number if admission.patient else "N/A"
            ward_name = admission.ward.name if admission.ward else "N/A"
            bed_number = admission.bed.bed_number if admission.bed else "N/A"
            admission_date = admission.admission_date.strftime('%Y-%m-%d') if admission.admission_date else 'N/A'
            discharge_date = admission.discharge_date.strftime('%Y-%m-%d') if admission.discharge_date else 'N/A'
            status = admission.status.value.title() if hasattr(admission.status, 'value') else str(admission.status).title()
            los = admission_los.get(admission.id, 0)
            revenue = float(admission.invoice.total_amount) if admission.invoice else 0.00
            encounters_count = len(admission.encounters) if admission.encounters else 0
            
            ws.cell(row=row, column=1, value=admission.admission_number).border = border
            ws.cell(row=row, column=2, value=patient_name).border = border
            ws.cell(row=row, column=3, value=patient_number).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=ward_name).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=bed_number).border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=admission_date).border = border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=7, value=discharge_date).border = border
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=8, value=status).border = border
            ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=9, value=los).border = border
            ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=10, value=revenue).border = border
            ws.cell(row=row, column=10).number_format = '#,##0.00'
            ws.cell(row=row, column=10).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=11, value=encounters_count).border = border
            ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_opd_report_csv(context: Dict[str, Any]) -> str:
    """Fallback CSV for OPD report."""
    from app.utils.csv_generator import generate_opd_report_csv
    return generate_opd_report_csv(context)


def _generate_ipd_report_csv(context: Dict[str, Any]) -> str:
    """Fallback CSV for IPD report."""
    from app.utils.csv_generator import generate_ipd_report_csv
    return generate_ipd_report_csv(context)


def generate_disease_report_excel(context: Dict[str, Any]) -> bytes:
    """Generate detailed Excel report for disease encounter statistics."""
    if not OPENPYXL_AVAILABLE:
        from app.utils.csv_generator import generate_disease_report_csv
        return generate_disease_report_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Disease Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    # Title
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1, value="DISEASE ENCOUNTER REPORT").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context['start_date']} to {context['end_date']}")
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%Y-%m-%d %H:%M'))
    row += 1
    ws.cell(row=row, column=1, value="Total Encounters:").font = subtitle_font
    ws.cell(row=row, column=2, value=context.get('total_encounters', 0))
    row += 2
    
    # Gender Distribution
    gender_breakdown = context.get('gender_breakdown', {})
    if gender_breakdown:
        ws.cell(row=row, column=1, value="GENDER DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Gender', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(gender_breakdown.values())
        for gender, count in gender_breakdown.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            ws.cell(row=row, column=1, value=gender or 'Unknown').border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Age Distribution
    age_groups = context.get('age_groups', {})
    if age_groups:
        ws.cell(row=row, column=1, value="AGE DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Age Group', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(age_groups.values())
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            age_label = f"{age_group} years" if age_group != "65+" else "65+ years"
            ws.cell(row=row, column=1, value=age_label).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Disease Breakdown
    stats = context.get('stats', [])
    top_diseases = context.get('top_diseases', [])
    disease_counts = context.get('disease_counts', {})
    
    if stats:
        ws.cell(row=row, column=1, value="DISEASE ENCOUNTER BREAKDOWN").font = subtitle_font
        row += 1
        headers = ['Disease', 'Code', 'Total Encounters', 'Primary Count', 'Primary %', 'First Recorded', 'Last Recorded']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [2, 3, 4, 5] else 'left' if col == 1 else 'center')
        row += 1
        
        for item in stats:
            first_recorded = item.get('first_recorded')
            last_recorded = item.get('last_recorded')
            ws.cell(row=row, column=1, value=item.get('name', 'N/A')).border = border
            ws.cell(row=row, column=2, value=item.get('code', '—')).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=item.get('encounter_count', 0)).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=item.get('primary_count', 0)).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=f"{item.get('primary_ratio', 0):.1f}%").border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=first_recorded.strftime('%Y-%m-%d') if first_recorded else 'N/A').border = border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=7, value=last_recorded.strftime('%Y-%m-%d') if last_recorded else 'N/A').border = border
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
            row += 1
    elif top_diseases:
        ws.cell(row=row, column=1, value="TOP DISEASES").font = subtitle_font
        row += 1
        headers = ['Disease', 'Encounter Count', 'Unique Patients']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        row += 1
        
        disease_patient_counts = context.get('disease_patient_counts', {})
        for disease_name, count in top_diseases:
            patient_count = disease_patient_counts.get(disease_name, 0)
            ws.cell(row=row, column=1, value=disease_name).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=patient_count).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_financial_report_excel(context: Dict[str, Any]) -> bytes:
    """Generate detailed Excel report for financial report."""
    if not OPENPYXL_AVAILABLE:
        from app.utils.csv_generator import generate_financial_report_csv
        return generate_financial_report_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:I{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:I{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:I{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    # Title
    ws.merge_cells(f'A{row}:I{row}')
    ws.cell(row=row, column=1, value="FINANCIAL REPORT").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context['start_date']} to {context['end_date']}")
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%Y-%m-%d %H:%M'))
    row += 2
    
    # Summary Statistics
    ws.cell(row=row, column=1, value="SUMMARY STATISTICS").font = subtitle_font
    row += 1
    summary_data = [
        ['Total Revenue', f"GHS {float(context.get('total_revenue', Decimal('0.00'))):,.2f}"],
        ['Total Paid', f"GHS {float(context.get('total_paid', Decimal('0.00'))):,.2f}"],
        ['Outstanding Balance', f"GHS {float(context.get('total_outstanding', Decimal('0.00'))):,.2f}"],
        ['Total Invoices', len(context.get('invoices', []))],
        ['Total Charges', len(context.get('charges', []))],
        ['Total Payments', len(context.get('payments', []))],
    ]
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1
    
    # Service Type Breakdown
    service_breakdown = context.get('service_breakdown', {})
    if service_breakdown:
        ws.cell(row=row, column=1, value="REVENUE BY SERVICE TYPE").font = subtitle_font
        row += 1
        headers = ['Service Type', 'Count', 'Total Revenue (GHS)', 'Paid (GHS)', 'Outstanding (GHS)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        for service, data in sorted(service_breakdown.items(), key=lambda x: x[1]['total'], reverse=True):
            outstanding = data['total'] - data['paid']
            ws.cell(row=row, column=1, value=service.replace('_', ' ').title()).border = border
            ws.cell(row=row, column=2, value=data['count']).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=float(data['total'])).border = border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4, value=float(data['paid'])).border = border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5, value=float(outstanding)).border = border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Payment Mechanism Breakdown
    payment_breakdown = context.get('payment_breakdown', {})
    if payment_breakdown:
        ws.cell(row=row, column=1, value="REVENUE BY PAYMENT MECHANISM").font = subtitle_font
        row += 1
        headers = ['Payment Mechanism', 'Count', 'Total Revenue (GHS)', 'Paid (GHS)', 'Outstanding (GHS)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        for mechanism, data in sorted(payment_breakdown.items(), key=lambda x: x[1]['total'], reverse=True):
            outstanding = data['total'] - data['paid']
            ws.cell(row=row, column=1, value=mechanism.replace('_', ' ').title()).border = border
            ws.cell(row=row, column=2, value=data['count']).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=float(data['total'])).border = border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4, value=float(data['paid'])).border = border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5, value=float(outstanding)).border = border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Detailed Invoices
    invoices_detailed = context.get('invoices_detailed', [])
    if invoices_detailed:
        ws.cell(row=row, column=1, value="DETAILED INVOICES").font = subtitle_font
        row += 1
        headers = ['Invoice #', 'Date', 'Patient Name', 'Patient #', 'Service Types', 'Total (GHS)', 'Paid (GHS)', 'Balance (GHS)', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [2, 4] else 'left' if col in [1, 3, 5, 9] else 'right')
        row += 1
        
        for inv_detail in invoices_detailed:
            invoice = inv_detail['invoice']
            patient = inv_detail.get('patient')
            patient_name = f"{patient.first_name} {patient.last_name}" if patient else "N/A"
            patient_number = patient.patient_number if patient else "N/A"
            
            # Get service types from charges
            charges = inv_detail.get('charges', [])
            service_types = [c.charge_type.value.replace('_', ' ').title() for c in charges]
            service_type_str = ', '.join(service_types) if service_types else 'N/A'
            
            invoice_date = invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else 'N/A'
            total = float(invoice.total_amount)
            paid = float(invoice.paid_amount or Decimal('0.00'))
            balance = float(invoice.balance or Decimal('0.00'))
            status = "Paid" if balance <= 0 else "Partial" if paid > 0 else "Unpaid"
            
            ws.cell(row=row, column=1, value=invoice.invoice_number or f"#{invoice.id}").border = border
            ws.cell(row=row, column=2, value=invoice_date).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=patient_name).border = border
            ws.cell(row=row, column=4, value=patient_number).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=service_type_str).border = border
            ws.cell(row=row, column=6, value=total).border = border
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=7, value=paid).border = border
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=8, value=balance).border = border
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=8).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=9, value=status).border = border
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_pharmacy_report_excel(context: Dict[str, Any]) -> bytes:
    """Generate detailed Excel report for pharmacy report."""
    if not OPENPYXL_AVAILABLE:
        from app.utils.csv_generator import generate_pharmacy_report_csv
        return generate_pharmacy_report_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pharmacy Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    # Title
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="PHARMACY REPORT").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context['start_date']} to {context['end_date']}")
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%Y-%m-%d %H:%M'))
    row += 1
    ws.cell(row=row, column=1, value="Total Prescriptions:").font = subtitle_font
    ws.cell(row=row, column=2, value=context.get('total_prescriptions', 0))
    row += 1
    ws.cell(row=row, column=1, value="Total Revenue:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"GHS {float(context.get('total_revenue', Decimal('0.00'))):,.2f}")
    row += 2
    
    # Top Medications
    top_medications = context.get('top_medications', [])
    if top_medications:
        ws.cell(row=row, column=1, value="TOP PRESCRIBED MEDICATIONS").font = subtitle_font
        row += 1
        headers = ['Rank', 'Medication', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 1 or col == 3 else 'left')
        row += 1
        
        for idx, (med, count) in enumerate(top_medications[:20], 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=med).border = border
            ws.cell(row=row, column=3, value=count).border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            row += 1
        row += 1
    
    # Gender Distribution
    gender_breakdown = context.get('gender_breakdown', {})
    if gender_breakdown:
        ws.cell(row=row, column=1, value="GENDER DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Gender', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(gender_breakdown.values())
        for gender, count in gender_breakdown.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            ws.cell(row=row, column=1, value=gender or 'Unknown').border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Age Distribution
    age_groups = context.get('age_groups', {})
    if age_groups:
        ws.cell(row=row, column=1, value="AGE DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Age Group', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(age_groups.values())
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            age_label = f"{age_group} years" if age_group != "65+" else "65+ years"
            ws.cell(row=row, column=1, value=age_label).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_lab_report_excel(context: Dict[str, Any]) -> bytes:
    """Generate detailed Excel report for lab report."""
    if not OPENPYXL_AVAILABLE:
        from app.utils.csv_generator import generate_lab_report_csv
        return generate_lab_report_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    # Title
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="LAB REPORT").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context['start_date']} to {context['end_date']}")
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%Y-%m-%d %H:%M'))
    row += 1
    ws.cell(row=row, column=1, value="Total Orders:").font = subtitle_font
    ws.cell(row=row, column=2, value=context.get('total_orders', 0))
    row += 1
    ws.cell(row=row, column=1, value="Total Revenue:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"GHS {float(context.get('total_revenue', Decimal('0.00'))):,.2f}")
    row += 2
    
    # Test Frequency
    test_counts = context.get('test_counts', {})
    if test_counts:
        ws.cell(row=row, column=1, value="TEST FREQUENCY").font = subtitle_font
        row += 1
        headers = ['Test Name', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left')
        row += 1
        
        for test_name, count in sorted(test_counts.items(), key=lambda x: x[1], reverse=True)[:20]:
            ws.cell(row=row, column=1, value=test_name).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            row += 1
        row += 1
    
    # Gender Distribution
    gender_breakdown = context.get('gender_breakdown', {})
    if gender_breakdown:
        ws.cell(row=row, column=1, value="GENDER DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Gender', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(gender_breakdown.values())
        for gender, count in gender_breakdown.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            ws.cell(row=row, column=1, value=gender or 'Unknown').border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Age Distribution
    age_groups = context.get('age_groups', {})
    if age_groups:
        ws.cell(row=row, column=1, value="AGE DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Age Group', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(age_groups.values())
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            age_label = f"{age_group} years" if age_group != "65+" else "65+ years"
            ws.cell(row=row, column=1, value=age_label).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_radiology_report_excel(context: Dict[str, Any]) -> bytes:
    """Generate detailed Excel report for radiology report."""
    if not OPENPYXL_AVAILABLE:
        from app.utils.csv_generator import generate_radiology_report_csv
        return generate_radiology_report_csv(context).encode('utf-8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Radiology Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Hospital header
    hospital_settings = context.get('hospital_settings')
    if hospital_settings:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1, value=hospital_settings.hospital_name or "Hospital").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        if hospital_settings.hospital_address:
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row=row, column=1, value=hospital_settings.hospital_address)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
    else:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1, value="HOSPITAL").font = title_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
    
    # Title
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="RADIOLOGY REPORT").font = title_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    row += 2
    
    # Period
    ws.cell(row=row, column=1, value="Period:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"{context['start_date']} to {context['end_date']}")
    row += 1
    report_date = context.get('report_date', datetime.now())
    if isinstance(report_date, datetime):
        ws.cell(row=row, column=1, value="Report Generated:").font = subtitle_font
        ws.cell(row=row, column=2, value=report_date.strftime('%Y-%m-%d %H:%M'))
    row += 1
    ws.cell(row=row, column=1, value="Total Orders:").font = subtitle_font
    ws.cell(row=row, column=2, value=context.get('total_orders', 0))
    row += 1
    ws.cell(row=row, column=1, value="Total Revenue:").font = subtitle_font
    ws.cell(row=row, column=2, value=f"GHS {float(context.get('total_revenue', Decimal('0.00'))):,.2f}")
    row += 2
    
    # Study Type Frequency
    study_counts = context.get('study_counts', {})
    if study_counts:
        ws.cell(row=row, column=1, value="STUDY TYPE FREQUENCY").font = subtitle_font
        row += 1
        headers = ['Study Type', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left')
        row += 1
        
        for study_type, count in sorted(study_counts.items(), key=lambda x: x[1], reverse=True)[:20]:
            ws.cell(row=row, column=1, value=study_type).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            row += 1
        row += 1
    
    # Gender Distribution
    gender_breakdown = context.get('gender_breakdown', {})
    if gender_breakdown:
        ws.cell(row=row, column=1, value="GENDER DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Gender', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(gender_breakdown.values())
        for gender, count in gender_breakdown.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            ws.cell(row=row, column=1, value=gender or 'Unknown').border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
        row += 1
    
    # Age Distribution
    age_groups = context.get('age_groups', {})
    if age_groups:
        ws.cell(row=row, column=1, value="AGE DISTRIBUTION").font = subtitle_font
        row += 1
        headers = ['Age Group', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 2 else 'left' if col == 1 else 'right')
        row += 1
        
        total_patients = sum(age_groups.values())
        for age_group, count in age_groups.items():
            percentage = (count / total_patients * 100) if total_patients > 0 else 0
            age_label = f"{age_group} years" if age_group != "65+" else "65+ years"
            ws.cell(row=row, column=1, value=age_label).border = border
            ws.cell(row=row, column=2, value=count).border = border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
